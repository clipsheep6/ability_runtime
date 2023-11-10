#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Copyright (c) 2023 Huawei Device Co., Ltd.
Licensed under the Apache License, Version 2.0 (the "License");
you may not use this file except in compliance with the License.
You may obtain a copy of the License at

    http://www.apache.org/licenses/LICENSE-2.0

Unless required by applicable law or agreed to in writing, software
distributed under the License is distributed on an "AS IS" BASIS,
WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
See the License for the specific language governing permissions and
limitations under the License.

Description: Use ark to execute workload test suite
"""

import argparse
import os
import subprocess
import platform
import errno
import sys
import importlib
import glob
import datetime
import stat
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill


class WorkLoadConfig:
    TEST_WORKLOAD_GIT_URL = 'https://gitee.com/xliu-huanwei/ark-workload.git'
    TEST_WORKLOAD_GIT_HASH = '90236fa3aa853db7af56d80cc6391432a51a1601'
    START_INDEX = -19
    END_INDEX = -5


def str_to_bool(value):
    if isinstance(value, bool):
        return value
    if value.lower() in ('true', '1'):
        return True
    elif value.lower() in ('false', '0'):
        return False
    else:
        raise argparse.ArgumentTypeError('Invalid boolean value: {}'.format(value))


def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument('--code-path', metavar='DIR',
                        help='code root path')
    parser.add_argument('--run-aot', default=False, nargs='?', type=str_to_bool,
                        help='Default Run run_pgo.sh, run-aot is true run run_aot.sh')
    parser.add_argument('--report', default=False, nargs='?', type=str_to_bool,
                        help='Support daily care performance results and script '
                             'preliminary analysis of performance data.')
    parser.add_argument('--tools-type', default='dev', nargs='?', help='tools type')
    parser.add_argument('--boundary-value', default=-10, nargs='?',
                        help='inferior boundary value')
    parser.add_argument('--run-count', default='10', nargs='?',
                        help='Compile all cases, execute the case count')
    return parser.parse_args()


def execute_shell_command(command: str):
    process = subprocess.Popen(command, shell=False)
    process.wait()


def execute_shell_command_add(command: list):
    for file in glob.glob(command[-1]):
        command[-1] = file
        process = subprocess.Popen(command, shell=False)
        process.wait()


def git_clone(repository_url: str, destination_path: str):
    command = ['git', 'clone', repository_url, destination_path]
    if platform.system() == "Windows":
        subprocess.run(command, check=True, shell=False)
    else:
        subprocess.run(command, check=True)


def git_checkout(commit_hash: str, destination_path: str):
    command = ['git', 'checkout', commit_hash]
    subprocess.run(command, cwd=destination_path)


def git_pull(check_out_dir=os.getcwd()):
    cmds = ['git', 'pull', '--rebase']
    with subprocess.Popen(cmds, cwd=check_out_dir) as proc:
        proc.wait()


def git_clean(clean_dir=os.getcwd()):
    cmds = ['git', 'checkout', '--', '.']
    with subprocess.Popen(cmds, cwd=clean_dir) as proc:
        proc.wait()


def execute_shell_script(script_path, args):
    command = ['sh', script_path] + args
    process = subprocess.Popen(command, stdout=subprocess.PIPE,
                               stderr=subprocess.PIPE, universal_newlines=True)
    while True:
        try:
            text_data = process.stdout.readline()
            sys.stdout.flush()
            if len(text_data.strip()) != 0:
                print(text_data.strip())
        except OSError as error:
            if error == errno.ENOENT:
                print("no such file")
            elif error == errno.EPERM:
                print("permission denied")
            break
        if not text_data:
            break
    process.wait()
    return_code = process.returncode
    if return_code != 0:
        error_output = process.stderr.read().strip()
        if error_output:
            print(error_output)


def configure_environment(path, file_path, tools_type):
    text = "--case-path ts-swift-workload\n" \
           f"--ts-tools-path {path}\n" \
           f"--tools-type {tools_type}\n" \
           "--swift-tools-path ~/tools/swift-5.7.3-RELEASE-ubuntu22.04/usr/bin\n" \
           "--android-ndk ~/apple/android-ndk-r25c\n" \
           "--Ninja-ReleaseAssert ~/apple/build/Ninja-ReleaseAssert\n" \
           "end"
    args = os.O_RDWR | os.O_CREAT
    file_descriptor = os.open(file_path, args, stat.S_IRUSR | stat.S_IWUSR)
    file_object = os.fdopen(file_descriptor, "w+")
    file_object.write(text)


def write_to_txt(file_path, text):
    args = os.O_RDWR | os.O_CREAT | os.O_APPEND
    file_descriptor = os.open(file_path, args, stat.S_IRUSR | stat.S_IWUSR)
    file_object = os.fdopen(file_descriptor, "w+")
    file_object.write(text)


def prepare_workload_code(path):
    data_dir = os.path.join("arkcompiler/ets_runtime/test/workloadtest/", "data")
    if path:
        data_dir = os.path.join(path, data_dir)
    if not os.path.isdir(os.path.join(data_dir, '.git')):
        git_clone(WorkLoadConfig.TEST_WORKLOAD_GIT_URL, data_dir)
        os.chdir(data_dir)
        git_checkout(WorkLoadConfig.TEST_WORKLOAD_GIT_HASH, './')
    else:
        os.chdir(data_dir)
        git_clean(data_dir)
        git_pull(data_dir)
        git_checkout(WorkLoadConfig.TEST_WORKLOAD_GIT_HASH, './')
    execute_shell_command_add(['chmod', '+x', '*.sh'])
    execute_shell_command_add(['chmod', '+x', '*.py'])
    try:
        importlib.import_module('openpyxl')
    except ImportError:
        execute_shell_command(['pip', 'install', 'openpyxl'])


def report(boundary_value: int):
    del_out_file()
    folder_path = "./"
    file_pattern = "pgo_data_*.xlsx"
    file_paths = glob.glob(os.path.join(folder_path, file_pattern))
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    file_paths.sort(key=lambda x: datetime.datetime.strptime(
        x[WorkLoadConfig.START_INDEX:WorkLoadConfig.END_INDEX],
        "%Y%m%d%H%M%S"), reverse=True)
    max_two_files = file_paths[:2]
    if len(max_two_files) == 2:
        wb_one = load_workbook(max_two_files[0])
        sheet_one = wb_one.active
        wb_two = load_workbook(max_two_files[1])
        sheet_two = wb_two.active
        data_one = []
        data_two = []
        for row in sheet_one.iter_rows(min_row=2, values_only=True):
            data_one.append(row)
        for row in sheet_two.iter_rows(min_row=2, values_only=True):
            data_two.append(row)
        if len(data_one) != len(data_two):
            return
        print('generate report dependent files:', max_two_files)
        result_data = []
        write_to_txt('../out/pgo_daily.txt', 'case:percentage\n')
        for row_one, row_two in zip(data_one, data_two):
            case = row_one[0]
            average_one = row_one[-1]
            average_two = row_two[-1]
            difference = (average_one - average_two) / average_one * 100
            percentage = "{:.2f}%".format(difference)
            result_data.append([case, percentage])
            write_to_txt('../out/pgo_daily.txt', case + percentage + '\n')
        result_wb = Workbook()
        result_sheet = result_wb.active
        result_sheet.append(['case', 'percentage'])
        for row in result_data:
            result_sheet.append(row)
            cell = result_sheet.cell(row=result_sheet.max_row, column=2)
            if cell.value and float(cell.value.strip('%')) < boundary_value:
                cell.fill = red_fill
        now = datetime.datetime.now()
        formatted_date = now.strftime("%Y%m%d%H%M%S")
        daily_excel_name = "pgo_daily_" + formatted_date + ".xlsx"
        daily_excel_out = "../out/pgo_daily.xlsx"
        result_wb.save(daily_excel_name)
        result_wb.save(daily_excel_out)


def del_out_file():
    destination_dir = '../out/'
    os.makedirs(destination_dir, exist_ok=True)
    file_list = os.listdir(destination_dir)
    for file_name in file_list:
        file_path = os.path.join(destination_dir, file_name)
        if os.path.isfile(file_path):
            os.remove(file_path)


def main(args):
    print("\nWait a moment..........\n")
    start_time = datetime.datetime.now()
    prepare_workload_code(args.code_path)
    tools_type = 'dev'
    if args.tools_type:
        tools_type = args.tools_type
    boundary_value = -10
    if args.boundary_value:
        boundary_value = args.boundary_value
    run_count = '10'
    if args.run_count:
        run_count = args.run_count
    if args.run_aot:
        print('execute run_aot.sh is currently not supported')
    else:
        configure_environment(args.code_path, 'toolspath.txt', tools_type)
        execute_args = ['--build', '--excel', '--run-count']
        execute_args.append(run_count)
        execute_shell_script("run_pgo.sh", execute_args)
    end_time = datetime.datetime.now()
    print(f"used time is: {str(end_time - start_time)}")
    if args.report:
        try:
            report(int(args.boundary_value))
        except ValueError:
            print('args.boundary_value value should be a number')


if __name__ == "__main__":
    sys.exit(main(parse_args()))
